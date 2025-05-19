"""
Telegram Бот: Рандомайзер рациона "Ложка_бот"
Автор: Enzhe Akhmetova
GitHub: https://github.com/enaenaenahm
Год создания: 2025
"""
import logging
import random
import asyncio
import asyncpg
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters.command import Command
import re
import os
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile, CallbackQuery
from datetime import datetime
from collections import defaultdict
from aiogram.types import Message
from PIL import Image
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import traceback
from aiogram.utils.keyboard import InlineKeyboardBuilder

DB_CONFIG = {
    "user": "my_user",
    "password": "root",
    "database": "diet_bot_db",
    "host": "localhost",
    "port": "5432"
}

NAME_REGEX = re.compile(r'^[A-Za-zА-Яа-яЁё\s\-]+$')

logging.basicConfig(level=logging.INFO)
API_TOKEN = "8130078332:AAG5ISPKao2NGXD5TzrU1nNAdiIeZO_1wPs"

bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(storage=storage)

class Registration(StatesGroup):
    name = State()
    confirm = State()

class FeedbackForm(StatesGroup):
    waiting_for_message = State()

# tables in a database
async def init_db():
    pool = await asyncpg.create_pool(**DB_CONFIG)
    async with pool.acquire() as conn:
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id BIGINT PRIMARY KEY,
                name TEXT NOT NULL,
                registered_at TIMESTAMP DEFAULT NOW()
            )
        ''')

        await conn.execute('''
            CREATE TABLE IF NOT EXISTS recipes (
                recipe_id SERIAL PRIMARY KEY,
                meal_type TEXT NOT NULL,
                content TEXT NOT NULL,
                is_premium BOOLEAN DEFAULT FALSE
            )
        ''')

        await conn.execute('''
            CREATE TABLE IF NOT EXISTS subscriptions (
                subscription_id SERIAL PRIMARY KEY,
                user_id BIGINT REFERENCES users(user_id),
                start_date TIMESTAMP,
                end_date TIMESTAMP
            )
        ''')

        await conn.execute('''
            CREATE TABLE IF NOT EXISTS feedback (
                feedback_id SERIAL PRIMARY KEY,
                user_id BIGINT REFERENCES users(user_id),
                message TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS user_limits (
                user_id BIGINT PRIMARY KEY,
                weekly_used INT DEFAULT 0,
                subscribed BOOLEAN DEFAULT FALSE
            )
        ''')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS user_sent_recipes (
                user_id BIGINT,
                recipe_id INT,
                meal_type TEXT,
                PRIMARY KEY (user_id, recipe_id)
            )
        ''')
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS user_meal_limits (
                user_id BIGINT REFERENCES users(user_id),
                meal_type TEXT,
                used_count INTEGER DEFAULT 0,
                PRIMARY KEY (user_id, meal_type)
    )
        ''')
    await pool.close()

# Class for working with recipes
class RecipeCRUD:
    def __init__(self, pool):
        self.pool = pool

    async def get_recipes_by_meal_type(self, meal_type: str, include_premium: bool = False):
        async with self.pool.acquire() as conn:
            return await conn.fetch(
                "SELECT * FROM recipes WHERE meal_type = $1 AND (is_premium = FALSE OR $2 = TRUE)",
                meal_type, 
                include_premium
            )
    async def get_all_free_recipes(self):
        async with self.pool.acquire() as conn:
            return await conn.fetch('SELECT * FROM recipes WHERE is_premium = FALSE')

    async def get_free_recipes(self, meal_type: str):
        async with self.pool.acquire() as conn:
            return await conn.fetch('''
                SELECT * FROM recipes
                WHERE meal_type = $1 AND is_premium = FALSE
            ''', meal_type)

    async def get_recipes_by_meal(self, meal_type: str, subscribed: bool = False):
        query = "SELECT * FROM recipes WHERE meal_type = $1"
        if not subscribed:
            query += " AND is_premium = FALSE"

        async with self.pool.acquire() as conn:
            return await conn.fetch(query, meal_type)
    async def get_all_recipes(self, subscribed: bool = False):
        query = "SELECT * FROM recipes"
        if not subscribed:
            query += " WHERE is_premium = FALSE"

        async with self.pool.acquire() as conn:
            return await conn.fetch(query)


# Keyboards
async def main_kb(is_registered: bool):
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Рацион 🍽"), KeyboardButton(text="Услуги 💼")],
            [KeyboardButton(text="Обратная связь 📩")]
        ] if is_registered else [[KeyboardButton(text="Регистрация")]],
        resize_keyboard=True
    )


def services_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Подписка 🔓"), KeyboardButton(text="Сопровождение 👨‍🍳")],
            [KeyboardButton(text="Назад")]
        ],
        resize_keyboard=True
    )
def weekly_menu_options_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="Готовить каждый день", callback_data="week_daily"),
            InlineKeyboardButton(text="Готовить на 2-3 дня", callback_data="week_bulk")
        ]
    ])

def diet_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Завтрак"), KeyboardButton(text="Обед")],
            [KeyboardButton(text="Ужин"), KeyboardButton(text="Перекус")],
            [KeyboardButton(text="Меню на день"), KeyboardButton(text="Меню на неделю")],
            [KeyboardButton(text="Назад")]
        ],
        resize_keyboard=True
    )

@dp.message(Command("start"))
async def start(message: types.Message):
    user_name = message.from_user.first_name
    pool = await asyncpg.create_pool(**DB_CONFIG)
    async with pool.acquire() as conn:
        user = await conn.fetchrow(
            'SELECT * FROM users WHERE user_id = $1', message.from_user.id
        )
    await pool.close()
    is_registered = bool(user)
    if not is_registered:
        text = (
            "Этот бот для вас, если вы:\n"
        "🔹 устали от однообразной еды в своей тарелке\n"
        "🔹 ищете адекватную схему питания без ограничений\n"
        "🔹цените свое время и не хотите часами стоять на кухне\n"
        "и самое главное, устали задаваться вопросом \"что же мне сегодня поесть?\"🤔\n"
        "🥗Что может бот?\n"
        "Предложить случайный рецепт на выбранный прием пищи из базы или быстро и лично для вас составить сбалансированный и вкусный рацион из 4 приемов пищи на каждый день/целую неделю сразу 📋\n"
        "Мы готовы забрать все заботы о вашем питании на себя!\n"
        "Сделаем его разнообразным, вкусным и полезным! Жми кнопку Регистрация 🫶"
            )
    else:
        user_name = user['name']
        text = (
            "Привет! Меня зовут Алия, я фитнес-тренер 🏋️‍♀️ и нутрициолог 🥦\n\n"
            f"Приятно познакомиться, {user_name}! 🫶\n"
            "Теперь ты с нами — а значит, путь к разнообразному и сбалансированному питанию стал короче 🚀\n"
            "Преимущества нашего бота в том, что рецепты составлены из простых и доступных продуктов, "
            "готовятся быстро и легко. А главное — каждый прием пищи сбалансирован, полезен и вкусен. "
            "Твое тело точно оценит полезность, насыщенность и разнообразие рациона. 🥗\n\n"
            "Вот что умеет наш бот:\n"
            "🍳 Подбирает рецепты под твои предпочтения — случайно, быстро, без мук выбора\n"
            "🧾 Составляет рацион на день или неделю: вкусно, сбалансировано, удобно\n"
            "🛒 Формирует список продуктов на основе подобранных рецептов — ничего лишнего, только нужное\n\n"
            "В бесплатной версии ты получишь доступ к ограниченному набору рецептов — по 3 на каждый прием пищи "
            "(завтрак, обед, ужин, перекус). Этого достаточно, чтобы попробовать и влюбиться 💛\n\n"
            "А вот что даст подписка:\n"
            "✨ Полный доступ к большой базе рецептов\n"
            "📅 Меню на день и неделю станет действительно разнообразным — больше рецептов, больше комбо, никакой повторяющейся еды\n\n"
            "Готов пробовать? Начни с бесплатной версии — уже сейчас ты можешь собрать *Рацион* и посмотреть, как это удобно!\n"
            "А когда захочешь больше — оформить подписку можно всего в два клика прямо в разделе *Услуги* 🙌"
        )
    await message.answer(text, reply_markup=await main_kb(is_registered))

# Registration
@dp.message(F.text == "Регистрация")
async def start_registration(message: types.Message, state: FSMContext):
    pool = await asyncpg.create_pool(**DB_CONFIG)
    async with pool.acquire() as conn:
        user = await conn.fetchrow('SELECT * FROM users WHERE user_id = $1', message.from_user.id)
    await pool.close()
    if user:
        await message.answer("Вы уже зарегистрированы!", reply_markup=await main_kb(message.from_user.id))
    else:
        await message.answer("Введите ваше имя:", reply_markup=types.ReplyKeyboardRemove())
        await state.set_state(Registration.name)


@dp.message(Registration.name)
async def process_name(message: types.Message, state: FSMContext):
    if not NAME_REGEX.fullmatch(message.text):
        await message.answer("❌ Имя может содержать только русские/английские буквы, пробелы и дефисы.\nПопробуйте еще раз:")
        return
    await state.update_data(name=message.text)
    await message.answer("Подтвердите согласие на обработку данных (Да/Нет):")
    await state.set_state(Registration.confirm)


@dp.message(Registration.confirm)
async def process_confirm(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if message.text.lower() == "да":
        data = await state.get_data()
        pool = await asyncpg.create_pool(**DB_CONFIG)
        async with pool.acquire() as conn:
            await conn.execute('''
                INSERT INTO users(user_id, name)
                VALUES($1, $2)
                ON CONFLICT (user_id) DO NOTHING
            ''', user_id, data['name'])
        await pool.close()
        await message.answer("✅ Регистрация завершена!", reply_markup=await main_kb(True))
    else:
        await message.answer("❌ Регистрация отменена", reply_markup=await main_kb(False))
    await state.clear()

@dp.message(F.text == "Услуги 💼")
async def services_menu(message: types.Message):
    await message.answer("Выберите услугу:", reply_markup=services_kb())

@dp.message(F.text == "Сопровождение 👨‍🍳")
async def support_service(message: types.Message):
    photo = FSInputFile("images/support.jpg")
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="Связаться с экспертом 💬", url="https://t.me/egorova_aliya")]
        ]
    )
    await message.answer_photo(
        photo=photo,
        caption="Привет! 👋\n\n"
            "Меня зовут Алия, я фитнес-тренер 🏋️‍♀️ и нутрициолог 🥦!\n\n"
            "Что такое личное сопровождение со мной:\n\n"
            "Сопровождение — это не просто \"откажись от сладкого 🍰, не ешь после 18:00 ⏰, иди побегай 🏃‍♀️ и тогда ты похудеешь\"."
            "Это полноценная работа над образом жизни 🌱.\n\n"
            "У меня нет готовых шаблонов или универсальных планов \"как получить результат\" 📉.\n"
            "С каждым человеком мы выстраиваем свой путь 🛤:\n"
            "🥗 рацион,\n"
            "🛒 набор продуктов,\n"
            "💪 тренировки,\n"
            "🛏 режим,\n"
            "🔁 привычки.\n"
            "Все эти шаги вводим в жизнь постепенно, поэтому результат достигается плавно, но остается навсегда 🔒.\n\n"
            "Работа со мной — это сбор моего опыта, знаний, секретиков, лайфхаков и фишек ✨.\n"
            "Чтобы путь к телу мечты и любви к себе прошёл комфортно 🤍\n",
        reply_markup=keyboard,
        parse_mode="Markdown"
    )

@dp.message(F.text == "Рацион 🍽")
async def show_diet_menu(message: types.Message):
    await message.answer("Выберите прием пищи:", reply_markup=diet_kb())

# resipes
@dp.message(F.text.in_({"Завтрак", "Обед", "Ужин", "Перекус"}))
async def random_recipe(message: types.Message):
    user_id = message.from_user.id
    meal_type = message.text.lower()
    meal_type_ru = message.text  # Для красивого отображения
    pool = await asyncpg.create_pool(**DB_CONFIG)
    async with pool.acquire() as conn:
        try:
            # Checking the user's subscription
            user = await conn.fetchrow(
                "SELECT subscribed FROM users WHERE user_id = $1", 
                user_id
            )        
            if user and user['subscribed']:
                # For subscribers - no restrictions
                recipes = await conn.fetch(
                    """
                    SELECT r.recipe_id, r.content
                    FROM recipes r
                    WHERE r.meal_type = $1
                    AND r.recipe_id NOT IN (
                        SELECT recipe_id FROM user_sent_recipes
                        WHERE user_id = $2 AND meal_type = $1
                    )
                    """,
                    meal_type, user_id
                )
            else:
                # Updating the usage counter
                result = await conn.fetchrow(
                    """
                    INSERT INTO user_meal_limits (user_id, meal_type, used_count)
                    VALUES ($1, $2, 1)
                    ON CONFLICT (user_id, meal_type) 
                    DO UPDATE SET used_count = user_meal_limits.used_count + 1
                    RETURNING used_count
                    """,
                    user_id, meal_type
                )            
                if result['used_count'] > 3:
                    await message.answer(
                        f"Бесплатные рецепты для '{meal_type_ru}' закончились!\n"
                        "Ты посмотрел все доступные рецепты в бесплатной версии. 🍳\n"
                        "Хочешь больше разнообразия?\n"
                        "В подписке - большая база рецептов, и каждый день будет непохож на предыдущий!\n"
                        "Оформи подписку всего за пару кликов и открой полный доступ!💫 \n",
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                            InlineKeyboardButton(text="Оформить подписку", callback_data="subscribe")
                        ]])
                    )
                    return
                # We only receive free recipes that have not been shown yet
                recipes = await conn.fetch(
                    """
                    SELECT r.recipe_id, r.content
                    FROM recipes r
                    WHERE r.meal_type = $1
                    AND r.is_premium = FALSE
                    AND r.recipe_id NOT IN (
                        SELECT recipe_id FROM user_sent_recipes
                        WHERE user_id = $2 AND meal_type = $1
                    )
                    """,
                    meal_type, user_id
                )
            if recipes:
                chosen = random.choice(recipes)
                recipe_text = chosen['content']
                recipe_id = chosen['recipe_id']
                await message.answer(f"🍴 {meal_type_ru}:\n{recipe_text}")
                await conn.execute(
                    "INSERT INTO user_sent_recipes (user_id, recipe_id, meal_type) VALUES ($1, $2, $3)",
                    user_id, recipe_id, meal_type
                )
            else:
                await message.answer("Рецепты не найдены")
        except Exception as e:
            logging.error(f"Error: {str(e)}")
            await message.answer("Произошла ошибка, попробуйте позже")       
    await pool.close()

@dp.callback_query(F.data == "subscribe")
async def handle_subscribe(callback: CallbackQuery):
    pool = await asyncpg.create_pool(**DB_CONFIG)
    async with pool.acquire() as conn:
        user_id = callback.from_user.id
        await conn.execute(
            "UPDATE users SET subscribed = TRUE WHERE user_id = $1",
            user_id
        )
        await conn.execute(
            """
            INSERT INTO user_limits (user_id, daily_used, weekly_used)
            VALUES ($1, 0, 0)
            ON CONFLICT (user_id)
            DO UPDATE SET daily_used = 0, weekly_used = 0
            """,
            user_id
        )
    await callback.message.answer(
        "Вы успешно оформили подписку! Теперь вам доступны *все* рецепты и неограниченное количество генераций меню. 🎉",
        parse_mode="Markdown"
    )
    await callback.answer()
        
@dp.message(F.text == "Меню на день")
async def daily_menu(message: Message, state: FSMContext):
    user_id = message.from_user.id
    pool = await asyncpg.create_pool(**DB_CONFIG)
    crud = RecipeCRUD(pool)
    
    try:
        async with pool.acquire() as conn:
            # Checking subscription from users table as in weekly_batch_menu
            subscribed = await conn.fetchval(
                "SELECT subscribed FROM users WHERE user_id = $1", 
                user_id
            ) or False

            # Limit logic for UNsubscribed only
            if not subscribed:
                # Check current limits from user_limits
                record = await conn.fetchrow(
                    "SELECT daily_used FROM user_limits WHERE user_id = $1",
                    user_id
                )
                
                # Initialization on first use
                if not record:
                    await conn.execute(
                        "INSERT INTO user_limits(user_id, daily_used) VALUES ($1, 0)",
                        user_id
                    )
                    remaining = 3
                else:
                    remaining = 3 - record['daily_used']

                # Blocking when limits are exhausted
                if remaining <= 0:
                    return await message.answer(
                        "Ты уже составил несколько дневных меню, но с таким небольшим количеством рецептов их хватит на несколько комбинаций.\n"
                        "Меню быстро надоест, а полноценного перехода на сбалансированное питание не получится.🥲\n"
                        "А ведь оно помогает подтянуть кожу, улучшить состояние лица, избавиться от отеков и почувствовать легкость в теле.💪\n"
                        "Хочешь реальных изменений?\n"
                        "Подпишись и открой доступ к большому количеству рецептов - всего пара кликов!💫\n",
                        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                            InlineKeyboardButton(text="Оформить подписку", callback_data="subscribe")
                        ]])
                    )
                # Counter update
                await conn.execute(
                    "UPDATE user_limits SET daily_used = daily_used + 1 WHERE user_id = $1",
                    user_id
                )
        # Receiving recipes with your subscription
        meals = ["завтрак", "обед", "перекус", "ужин"]
        all_recipes = {}
        for meal in meals:
            recipes = await crud.get_recipes_by_meal_type(
                meal_type=meal,
                include_premium=subscribed  # Показываем платные для подписанных
            )
            all_recipes[meal] = recipes or []
        await state.update_data(
            all_recipes=all_recipes,
            current_meal_index=0,
            current_recipe_index=0,
            selected_recipes={}
        )
        await show_current_recipe(message, state)     
    except Exception as e:
        logging.error(f"Error in daily_menu: {e}")
        await message.answer("Произошла ошибка при формировании меню. Попробуй позже 🛠️")
    finally:
        await pool.close()

async def show_current_recipe(message: Message, state: FSMContext):
    data = await state.get_data()
    all_recipes = data['all_recipes']
    current_meal_index = data['current_meal_index']
    current_recipe_index = data.get('current_recipe_index', 0)
    meals = list(all_recipes.keys())

    if current_meal_index >= len(meals):
        selected_recipes = data['selected_recipes']
        selected_text = ""
        for meal, recipes in selected_recipes.items():
            selected_text += f"\n\n{get_emoji(meal)} {meal.capitalize()}:\n"
            for recipe in recipes:
                selected_text += f"- {recipe['content']}\n"

        # Generate a shopping list
        shopping_list = await generate_shopping_list(selected_recipes)
        await message.answer(f"Ваше меню на день:\n{selected_text}")
        return await message.answer(f"🛒 Список покупок:\n{shopping_list}")
    meal = meals[current_meal_index]
    recipes = all_recipes[meal]
    if not recipes:
        recipe_text = "Нет доступных рецептов"
    else:
        recipe = recipes[current_recipe_index % len(recipes)]
        recipe_text = recipe['content']
    text = f"{get_emoji(meal)} {meal.capitalize()}:\n{recipe_text}"
    builder = InlineKeyboardBuilder()
    if recipes:
        builder.button(text="⬅️", callback_data="prev_recipe")
        builder.button(text="➡️", callback_data="next_recipe")

    builder.button(text="✅", callback_data="next_meal")
    await message.answer(text, reply_markup=builder.as_markup())

async def generate_shopping_list(selected_recipes: dict) -> str:
    shopping_items = {}
    for meal, recipes in selected_recipes.items():
        for recipe in recipes:
            ingredients = recipe.get("ingredients")
            if not ingredients:
                parsed = parse_recipe_content(recipe.get("content", ""))
                parsed_ingredients = parse_ingredients(parsed.get("ingredients", ""))
                ingredients = [
                    {
                        "name": name,
                        "quantity": amount,
                        "unit": unit
                    } for name, amount, unit in parsed_ingredients
                    if amount > 0  # ← фильтруем "по вкусу"
                ]
            for ingredient in ingredients:
                name = ingredient["name"]
                quantity = ingredient.get("quantity", 0)
                unit = ingredient.get("unit", "").lower()
                if quantity <= 0:
                    continue
                key = (name, unit)
                if key in shopping_items:
                    shopping_items[key] += quantity
                else:
                    shopping_items[key] = quantity
    shopping_list = ""
    for (name, unit), quantity in shopping_items.items():
        quantity = round(quantity, 2)
        unit_display = f" {unit}" if unit else ""
        shopping_list += f"• {name} – {quantity}{unit_display}\n"
    return shopping_list

def get_emoji(meal: str) -> str:
    emojis = {
        "завтрак": "🍳",
        "обед": "🍲",
        "ужин": "🥗",
        "перекус": "🥪"
    }
    return emojis.get(meal, "🍽")

# Daily Menu Button Handlers
@dp.callback_query(F.data == "next_recipe")
async def next_recipe(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    current_index = data.get("current_recipe_index", 0)
    await state.update_data(current_recipe_index=current_index + 1)
    await callback.message.delete()
    await show_current_recipe(callback.message, state)

@dp.callback_query(F.data == "prev_recipe")
async def prev_recipe(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    current_index = data.get("current_recipe_index", 0)
    await state.update_data(current_recipe_index=max(0, current_index - 1))
    await callback.message.delete()
    await show_current_recipe(callback.message, state)

@dp.callback_query(F.data == "next_meal")
async def next_meal(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    current_meal_index = data.get("current_meal_index", 0)
    current_recipe_index = data.get("current_recipe_index", 0)
    selected_recipes = data.get('selected_recipes', {})
    meal = list(data['all_recipes'].keys())[current_meal_index]
    recipes = data['all_recipes'][meal]
    if recipes:
        selected_recipes[meal] = selected_recipes.get(meal, [])
        selected_recipes[meal].append(recipes[current_recipe_index % len(recipes)])
    await state.update_data(
        current_meal_index=current_meal_index + 1,
        current_recipe_index=0,  
        selected_recipes=selected_recipes
    )
    await callback.message.delete()
    await show_current_recipe(callback.message, state)

@dp.message(F.text == "Меню на неделю")
async def choose_weekly_menu_mode(message: types.Message):
    await message.answer(
        "Как вы предпочитаете готовить на неделю?\nКаждый день готовить новые блюда или заготавливать на 2-3 дня?👇",
        reply_markup=weekly_menu_options_kb()
    )
# Menu Button Handlers for the Week
@dp.callback_query(F.data == "week_daily")
async def generate_weekly_menu_daily(callback: CallbackQuery):
    await handle_weekly_menu(callback, mode="daily")

@dp.callback_query(F.data == "week_bulk")
async def generate_weekly_menu_bulk(callback: CallbackQuery):
    await handle_weekly_menu(callback, mode="bulk")

@dp.message(F.text == "Готовить каждый день")
async def weekly_varied_menu(message: types.Message):
    user_id = message.from_user.id
    pool = await asyncpg.create_pool(**DB_CONFIG)
    crud = RecipeCRUD(pool)

    try:
        async with pool.acquire() as conn:
            # Subscription check
            subscribed = await conn.fetchval(
                "SELECT subscribed FROM users WHERE user_id = $1", 
                user_id
            ) or False

            if not subscribed:
                await message.answer(
                    "🥗 Эта функция доступна только для подписчиков.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(text="Оформить подписку", callback_data="subscribe")
                    ]])
                )
                return
            # Get recipes for every meal
            meals = ["завтрак", "обед", "перекус", "ужин"]
            recipes_by_meal_type = {}       
            for meal in meals:
                recipes = await crud.get_recipes_by_meal_type(
                    meal_type=meal,
                    include_premium=subscribed
                )
                recipes_by_meal_type[meal] = [r['content'] for r in recipes]

            # Excel
            excel_file = generate_weekly_excel_varied(recipes_by_meal_type)
            
            await message.answer_document(
                document=types.BufferedInputFile(
                    file=excel_file.getvalue(),
                    filename=f"varied_menu_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
                ),
                caption="📅 Меню: готовим каждый день"  
            )
    except Exception as e:
        logging.error(f"Ошибка: {e}")
        await message.answer("Произошла ошибка, попробуйте позже.")
    finally:
        await pool.close()

@dp.message(F.text == "Готовить на 2-3 дня")
async def weekly_batch_menu(message: types.Message):
    user_id = message.from_user.id
    pool = await asyncpg.create_pool(**DB_CONFIG)
    crud = RecipeCRUD(pool)
    try:
        async with pool.acquire() as conn:
            subscribed = await conn.fetchval(
                "SELECT subscribed FROM users WHERE user_id = $1", 
                user_id
            ) or False
            meals = ["завтрак", "обед", "перекус", "ужин"]
            recipes_by_meal_type = {}   
            for meal in meals:
                recipes = await crud.get_recipes_by_meal_type(
                    meal_type=meal,
                    include_premium=subscribed
                )
                recipes_by_meal_type[meal] = [r['content'] for r in recipes]

            # Excel
            excel_file = generate_bulk_excel(recipes_by_meal_type)
            
            await message.answer_document(
                document=types.BufferedInputFile(
                    file=excel_file.getvalue(),
                    filename=f"batch_menu_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
                ),
                caption="🧊 Меню: готовим на 2–3 дня"  
            )
    except Exception as e:
        logging.error(f"Ошибка: {e}")
        await message.answer("Произошла ошибка, попробуйте позже.")
    finally:
        await pool.close()
        # Excel
        logging.info(f"Start generating menu for user {user_id}")
        all_recipes = await crud.get_all_free_recipes()
        if not all_recipes:
            await message.answer("😔 В базе пока нет рецептов")
            return
        recipes_by_meal_type = {}
        meals = ["завтрак", "обед", "перекус", "ужин"]
        for meal in meals:
            recipes = await crud.get_recipes_by_meal_type(
                meal_type=meal, 
                include_premium=subscribed  # subscribed из предыдущей проверки
        )
        recipes_by_meal_type[meal] = [r['content'] for r in recipes]
        excel_file = generate_weekly_excel(recipes_by_meal_type)
        await message.answer_document(
            document=types.BufferedInputFile(
                file=excel_file.getvalue(),
                filename=f"weekly_menu_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            ),
            caption="Ваше меню на неделю 💫"  
        )

def generate_weekly_excel(recipes_by_meal_type: dict):
    try:
        wb = Workbook()
        ws_menu = wb.active
        ws_menu.title = "Меню"
        ws_shopping = wb.create_sheet("Список покупок")
        headers = ["День", "Прием пищи", "Блюдо", "Ингредиенты", "Инструкция", "КБЖУ"]
        ws_menu.append(headers)
        bold_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws_menu.cell(row=1, column=col).font = bold_font
        days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        meals_order = ["завтрак", "обед", "перекус", "ужин"]
        day_totals = defaultdict(lambda: {'calories': 0, 'proteins': 0, 'fats': 0, 'carbs': 0})
        shopping_list = defaultdict(lambda: defaultdict(float))
        used_recipes = defaultdict(list)  # meal_type -> list of recently used recipe titles
        for day_idx, day in enumerate(days, start=2):
            day_row_start = ws_menu.max_row + 1
            for meal in meals_order:
                recipes = recipes_by_meal_type.get(meal, [])
                if not recipes:
                    continue
                try:
                    # We exclude repetitions for the last 2 days
                    recent = used_recipes[meal][-2:]
                    available_recipes = [r for r in recipes if parse_recipe_content(r)['title'] not in recent]
                    if not available_recipes:
                        available_recipes = recipes
                    recipe_content = random.choice(available_recipes)
                    parsed = parse_recipe_content(recipe_content)
                    used_recipes[meal].append(parsed['title'])
                    logging.debug(f"Parsed recipe content: {parsed}") 
                    if parsed['kbju']:
                        kbju = parsed['kbju'].split('/')
                        if len(kbju) == 4:
                            cals = int(kbju[0])
                            proteins = int(kbju[1])
                            fats = int(kbju[2])
                            carbs = int(kbju[3])
                            day_totals[day]['calories'] += cals
                            day_totals[day]['proteins'] += proteins
                            day_totals[day]['fats'] += fats
                            day_totals[day]['carbs'] += carbs
                    row = ws_menu.max_row + 1
                    ws_menu.cell(row=row, column=1, value=day)
                    ws_menu.cell(row=row, column=2, value=meal.capitalize())
                    ws_menu.cell(row=row, column=3, value=parsed['title'])
                    logging.debug(f"Ingredients: {parsed['ingredients']}")
                    logging.debug(f"Instructions: {parsed['instructions']}")
                    cell_ingredients = ws_menu.cell(row=row, column=4, value=parsed['ingredients'])
                    cell_ingredients.alignment = Alignment(wrap_text=True)
                    if parsed['instructions']:
                        cell_instructions = ws_menu.cell(row=row, column=5, value=parsed['instructions'])
                        cell_instructions.alignment = Alignment(wrap_text=True)
                    else:
                        logging.warning(f"[!] Нет инструкции для рецепта: {parsed['title']}")
                        ws_menu.cell(row=row, column=5, value="")
                    ws_menu.cell(row=row, column=6, value=parsed['kbju'])
                    if parsed['ingredients']:
                        for ingredient in parse_ingredients(parsed['ingredients']):
                            name, amount, unit = ingredient
                            shopping_list[(name.lower(), unit)]['amount'] += amount
                except Exception as e:
                    logging.error(f"Error processing {meal} for {day}: {str(e)}")
                    continue
            if day in day_totals:
                total = day_totals[day]
                total_kbju = f"{total['calories']}/{total['proteins']}/{total['fats']}/{total['carbs']}"
                ws_menu.append([day, "Итого КБЖУ за день", "", "", "", total_kbju])
                for col in range(1, 7):
                    ws_menu.cell(row=ws_menu.max_row, column=col).font = bold_font
                    ws_menu.cell(row=ws_menu.max_row, column=col).fill = PatternFill(start_color="FFD3D3D3", fill_type="solid")
        ws_shopping.append(["Ингредиент", "Количество", "Ед.изм."])
        for (name, unit), data in shopping_list.items():
            ws_shopping.append([name.capitalize(), round(data['amount'], 2), unit if unit else '-'])
        for col in range(1, 4):
            ws_shopping.cell(row=1, column=col).font = bold_font
        for ws in [ws_menu, ws_shopping]:
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logging.error(f"Excel generation failed: {str(e)}")
        raise

async def handle_weekly_menu(callback: CallbackQuery, mode: str):
    user_id = callback.from_user.id
    pool = await asyncpg.create_pool(**DB_CONFIG)
    crud = RecipeCRUD(pool)

    try:
        # Checking subscription and limits as before (taking into account mode == "bulk" only for subscribers)
        async with pool.acquire() as conn:
            subscribed = await conn.fetchval(
                "SELECT subscribed FROM users WHERE user_id = $1", user_id
            )
            if not subscribed and mode == "daily":
                await callback.message.answer(
                    "Готовить каждый день разные блюда доступно только подписчикам. 💎\nОформите premium-подписку, чтобы разблокировать!",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                        InlineKeyboardButton(text="Оформить подписку", callback_data="subscribe")
                    ]])
                )
                return
        meals = ["завтрак", "обед", "перекус", "ужин"]
        recipes_by_meal_type = {}
        for meal in meals:
            recipes = await crud.get_recipes_by_meal(meal_type=meal, subscribed=subscribed)
            recipes_by_meal_type[meal] = [r['content'] for r in recipes if 'content' in r]
        if not recipes_by_meal_type:
            await callback.message.answer("В базе пока нет рецептов 😔")
            return
        # Selecting an Excel Generator
        if mode == "daily":
            file = generate_weekly_excel(recipes_by_meal_type)
        else:
            file = generate_bulk_excel(recipes_by_meal_type)
        await callback.message.answer_document(
            document=types.BufferedInputFile(
                file=file.getvalue(),
                filename=f"weekly_menu_{mode}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            ),
            caption="Ваше меню на неделю готово! 💫"
        )
    except Exception as e:
        logging.error(f"Ошибка в handle_weekly_menu: {str(e)}")
        await callback.message.answer("Произошла ошибка при создании меню")
    finally:
        await pool.close()

def generate_bulk_excel(recipes_by_meal_type: dict):
    try:
        wb = Workbook()
        ws_menu = wb.active
        ws_menu.title = "Меню"
        ws_shopping = wb.create_sheet("Список покупок")

        headers = ["День", "Прием пищи", "Блюдо", "Ингредиенты", "Инструкция", "КБЖУ"]
        ws_menu.append(headers)

        bold_font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws_menu.cell(row=1, column=col).font = bold_font
        days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        meals_order = ["завтрак", "обед", "перекус", "ужин"]
        shopping_list = defaultdict(lambda: defaultdict(float))
        meal_cache = {} 
        kbju_totals = defaultdict(lambda: {'calories': 0, 'proteins': 0, 'fats': 0, 'carbs': 0})
        for i, day in enumerate(days, start=2):
            for meal in meals_order:
                recipes = recipes_by_meal_type.get(meal, [])
                if not recipes:
                    continue
                # We update the recipe every 2 days
                if i % 2 == 0 or meal not in meal_cache:
                    recipe_content = random.choice(recipes)
                    meal_cache[meal] = parse_recipe_content(recipe_content)
                parsed = meal_cache[meal]
                row = ws_menu.max_row + 1
                ws_menu.cell(row=row, column=1, value=day)
                ws_menu.cell(row=row, column=2, value=meal.capitalize())
                ws_menu.cell(row=row, column=3, value=parsed['title'])
                cell_ingredients = ws_menu.cell(row=row, column=4, value=parsed['ingredients'])
                cell_ingredients.alignment = Alignment(wrap_text=True)
                if parsed['instructions']:
                    cell_instructions = ws_menu.cell(row=row, column=5, value=parsed['instructions'])
                    cell_instructions.alignment = Alignment(wrap_text=True)
                else:
                    ws_menu.cell(row=row, column=5, value="")
                ws_menu.cell(row=row, column=6, value=parsed['kbju'])
                # Shopping list
                if parsed['ingredients']:
                    for name, amount, unit in parse_ingredients(parsed['ingredients']):
                        shopping_list[(name.lower(), unit)]['amount'] += amount
                # KBJU
                if parsed['kbju']:
                    kbju = parsed['kbju'].split('/')
                    if len(kbju) == 4:
                        kbju_totals[day]['calories'] += int(kbju[0])
                        kbju_totals[day]['proteins'] += int(kbju[1])
                        kbju_totals[day]['fats'] += int(kbju[2])
                        kbju_totals[day]['carbs'] += int(kbju[3])
            # KBJU per day
            if day in kbju_totals:
                total = kbju_totals[day]
                kbju_row = f"{total['calories']}/{total['proteins']}/{total['fats']}/{total['carbs']}"
                ws_menu.append([day, "Итого КБЖУ за день", "", "", "", kbju_row])
                for col in range(1, 7):
                    ws_menu.cell(row=ws_menu.max_row, column=col).font = bold_font
                    ws_menu.cell(row=ws_menu.max_row, column=col).fill = PatternFill(start_color="FFD3D3D3", fill_type="solid")
        # Shopping list
        ws_shopping.append(["Ингредиент", "Количество", "Ед.изм."])
        for (name, unit), data in shopping_list.items():
            ws_shopping.append([name.capitalize(), round(data['amount'], 2), unit if unit else '-'])
        for col in range(1, 4):
            ws_shopping.cell(row=1, column=col).font = bold_font
        for ws in [ws_menu, ws_shopping]:
            for column in ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    except Exception as e:
        logging.error(f"[!] Ошибка при генерации Excel (bulk): {str(e)}")
        raise

def parse_ingredients(ingredients_str: str) -> list:
    pattern = re.compile(
        r"^(?P<name>.*?)"                                   # Ingredient name
        r"(?:\s*[-–—]?\s*|\s+)"                             # Optional hyphen or just space
        r"(?P<amount>[\d.,/]+)?"                            # Quantity (optional)
        r"\s*(?P<unit>[а-яa-zёЁ.]+)?"                       # Unit of measurement (optional)
        r"(?:\s*\(.*?\))?"                                  # Comments in brackets (ignored)
        r"(?:\s*по вкусу)?$",                               # "to taste" (optional)
        re.IGNORECASE
    )
    ingredients = []
    for line in ingredients_str.split('\n'):
        line = line.strip()
        if not line:
            continue
        match = pattern.match(line)
        if match:
            name = match.group("name").strip().capitalize()
            amount_str = match.group("amount")
            unit = (match.group("unit") or '').strip()
            if not amount_str and unit not in ["шт", "пучок", "зубчик", "лист", "веточка"]:
                continue  # skip ingredients without quantity

            try:
                if amount_str:
                    amount = float(eval(amount_str.replace(',', '.')))
                else:
                    amount = 1.0
            except:
                amount = 0.0
            if amount == 0:
                continue  
            ingredients.append((name, amount, unit))
        else:
            logging.warning(f"Не распознан ингредиент: {line}")
    return ingredients

def parse_recipe_content(content: str) -> dict:
    result = {
        'title': 'Без названия',
        'ingredients': 'Ингредиенты не указаны',
        'instructions': '   ',
        'kbju': ''
    }
    try:
        lines = [line.strip() for line in content.split('\n')]
        if not lines or all(not line for line in lines):
            return result
        result['title'] = lines[0]
        # Search KBZhU
        kbju_pattern = re.compile(r'КБЖУ\s*.*?(\d+/\d+/\d+/\d+)', re.IGNORECASE)
        for line in lines:
            match = kbju_pattern.search(line)
            if match:
                result['kbju'] = match.group(1)
                break
        ingredients = []
        instructions = []
        current_section = None
        # Using regular expressions to find sections
        for line in lines[1:]:
            # Check if the string starts with "Ingredients" or "Preparation"
            if re.fullmatch(r'ингредиенты\s*:?', line, re.IGNORECASE):
                current_section = 'ingredients'
                continue
            elif re.fullmatch(r'приготовление\s*:?', line, re.IGNORECASE):
                current_section = 'instructions'
                continue
            elif re.search(r'кбжу', line, re.IGNORECASE):
                current_section = None
                continue
            # Processing section contents
            if current_section == 'ingredients' and line:
                ingredients.append(line.lstrip('•').strip())
            elif current_section == 'instructions' and line:
                instructions.append(line)
        # Save the results
        if ingredients:
            result['ingredients'] = '\n'.join(ingredients)
        if instructions:
            result['instructions'] = '\n'.join(instructions)
    except Exception as e:
        logging.error(f"Error parsing recipe: {str(e)}")
    return result
  
@dp.message(F.text == "Подписка 🔓")
async def subscription_info(message: types.Message):
    text = (
        "*Мы рады, что тебе понравился наш бот!*🙏\n"
        "На данный момент мы работаем над полной версией с подпиской за 990 рублей в месяц и в скором времени его запустим! 🚀\n"
        "А пока ты можете опробовать вкуснейшие рационами в тестовом режиме! 🥗\n"
        "Мы обязательно сообщим о запуске полной версии бота, и как первый тестировщик ты сможешь воспользоваться им со скидкой 20% по промокоду *ТЕСТБОТ*🎉\n"
    )
    await message.answer(text, parse_mode="Markdown")

@dp.message(F.text == "Назад")
async def go_back_to_main(message: types.Message):
    # Check if the user is registered
    pool = await asyncpg.create_pool(**DB_CONFIG)
    async with pool.acquire() as conn:
        user = await conn.fetchrow('SELECT * FROM users WHERE user_id = $1', message.from_user.id)
    await pool.close()
    await message.answer("Вы вернулись в главное меню", reply_markup=await main_kb(is_registered=bool(user)))

@dp.message(F.text == "Обратная связь 📩")
async def start_feedback(message: types.Message, state: FSMContext):
    await message.answer("Пожалуйста, напишите ваше сообщение, мы обязательно его рассмотрим.")
    await state.set_state(FeedbackForm.waiting_for_message)

@dp.message(FeedbackForm.waiting_for_message)
async def save_feedback(message: types.Message, state: FSMContext):
    pool = await asyncpg.create_pool(**DB_CONFIG)
    try:
        async with pool.acquire() as conn:
            await conn.execute(
                "INSERT INTO feedback (user_id, message, created_at) VALUES ($1, $2, $3)",
                message.from_user.id,
                message.text,
                datetime.now()
            )
        await message.answer("Спасибо! Ваше сообщение отправлено.")
    finally:
        await pool.close()
    await state.clear()
async def main():
    await init_db()
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())